#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 18-9-26 下午7:48
# @Author  : zhuzhengyi
# @File    : deal_excl_xlsx.py
# @Software: PyCharm

import openpyxl as opxl
from os import listdir

def test(path):
	wb = opxl.load_workbook(path)
	print(type(wb))
	# sheet = wb.get_sheet_names
	# print(sheet)
	sheet = wb.active
	print(sheet.cell(3,1).value)

def main(sourcepath,resultpath):
	wb_res = opxl.load_workbook(resultpath)
	sheet_res = wb_res.active
	filelist = listdir(sourcepath)
	row_res = 1
	for file in filelist:
		wb = opxl.load_workbook(sourcepath+file)
		sheet = wb.active
		if row_res ==1:
			row = 1
		else:
			row = 2
		while sheet.cell(row, 1).value is not None:
			for clumn in range(1,16):
				sheet_res.cell(row_res,clumn).value=sheet.cell(row, clumn).value
			row +=1
			row_res+=1
	wb_res.save(resultpath)

if __name__ == '__main__':
	testpath = '/home/zzy/TrainData/test.xlsx'
	filespath = '/home/zzy/TrainData/xlsxfiles/'
	resultpath = '/home/zzy/TrainData/new.xlsx'
	# test(testpath)
	main(filespath,resultpath)